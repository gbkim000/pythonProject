#%% 데이터베이스 활용(1) : sqlite3

import sys, os, sqlite3
# from login import LoginWindow
# from PyQt5.QtWidgets import QApplication

def dbInsert():
    '''
    isolation_level=None이 : (실습을 위해) 쿼리문을 실행하여 DB에 즉시 반영, 즉시 자동 커밋.
    conn.commit(커밋)은 “변경사항을 DB에 반영한다”는 뜻, 
    commit을 하지 않으면 수정 작업에 대한 기록을 메모리에 임시로 저장
    commit과 반대되는 개념으로 conn.rollback(롤백)이 있다. 이전 이력으로 되돌린다는 뜻. 
    '''
    
    os.chdir("C:\\Users\\Administrator\\PyQtTest")      # 현재 디렉토리 변경        
    #print(os.getcwd())
    if os.path.exists("./test.db"):
        print("The 'test.db' file already exsists in this forder, That'll remove")       
        os.remove('test.db')
    
    conn = sqlite3.connect("test.db", isolation_level=None)
    
    # 커서 획득, 파이썬에서 파일을 읽고 쓰려면 커서를 가져와야 한다
    c = conn.cursor()
    
    # 테이블 생성 (데이터 타입은 TEST, NUMERIC, INTEGER, REAL, BLOB 등)
    c.execute("CREATE TABLE IF NOT EXISTS table1 (id integer PRIMARY KEY, name text, birthday text)")

    # 데이터 삽입 방법 1
    c.execute("INSERT INTO table1 VALUES(1, 'LEE', '1987-00-00')")
    
    # test_tuple = (
    #     (3, 'PARK', '1991-00-00'),
    #     (4, 'CHOI', '1999-00-00'),
    #     (5, 'JUNG', '1989-00-00'),
    #     (6, 'JANG', '1989-00-00')
    # )
    
    test_tuple = ((3, 'PARK', '1991-00-00'),)

    
    c.executemany("INSERT INTO table1(id, name, birthday) VALUES(?,?,?)", test_tuple)    
    #conn.commit()
    conn.close()
    
def dbRead():
    conn = sqlite3.connect("test.db", isolation_level=None)
    # 커서 획득, 파이썬에서 파일을 읽고 쓰려면 커서를 가져와야 한다
    c = conn.cursor()
    c.execute("SELECT * FROM table1")
    print(c.fetchone())
    print(c.fetchone())
    
    c.execute("SELECT * FROM table1")
    print(c.fetchall())
        
    # 방법 1
    c.execute("SELECT * FROM table1")
    for row in c.fetchall():
        print(row)
        
    # 방법 2
    for row in c.execute("SELECT * FROM table1 ORDER BY id ASC"):
        print(row)
    
    conn.close()
        
def dbSearch():
    conn = sqlite3.connect("test.db", isolation_level=None)
    # 커서 획득, 파이썬에서 파일을 읽고 쓰려면 커서를 가져와야 한다
    c = conn.cursor()

    # 방법 1
    param1 = (1,)
    c.execute('SELECT * FROM table1 WHERE id=?', param1)
    print('param1', c.fetchone())
    print('param1', c.fetchall())
    
    # 방법 2
    param2 = 1
    c.execute("SELECT * FROM table1 WHERE id='%s'" % param2)  # %s %d %f
    print('param2', c.fetchone())
    print('param2', c.fetchall())
    
    # 방법 3
    c.execute("SELECT * FROM table1 WHERE id=:Id", {"Id": 1})
    print('param3', c.fetchone())
    print('param3', c.fetchall())
    
    # 방법 4
    param4 = (1, 4)
    c.execute('SELECT * FROM table1 WHERE id IN(?,?)', param4)
    print('param4', c.fetchall())
    
    # 방법 5
    c.execute("SELECT * FROM table1 WHERE id In('%d','%d')" % (1, 4))
    print('param5', c.fetchall())
    
    # 방법 6
    c.execute("SELECT * FROM table1 WHERE id=:id1 OR id=:id2", {"id1": 1, "id2": 4})
    print('param6', c.fetchall())
        
    conn.close()
        
def dbUpdate():
    conn = sqlite3.connect("test.db", isolation_level=None)
    # 커서 획득, 파이썬에서 파일을 읽고 쓰려면 커서를 가져와야 한다
    c = conn.cursor()
    
    # 방법 1
    c.execute("UPDATE table1 SET name=? WHERE id=?", ('NEW1', 1))
    
    # 방법 2
    sql="UPDATE table1 SET name='%s' WHERE id='%s' " % ('NEW2', 3)
    print(sql)
    c.execute(sql)
    #c.execute("UPDATE table1 SET name=:name WHERE id=:id", {"name": 'NEW2', 'id': 3})

    # 방법 3
    c.execute("UPDATE table1 SET name='%s' WHERE id='%s'" % ('NEW3', 5))
    # 확인
    for row in c.execute('SELECT * FROM table1'):
        print(row)

    conn.close()
            
def dbDelete():   
    conn = sqlite3.connect("test.db", isolation_level=None)
    # 커서 획득, 파이썬에서 파일을 읽고 쓰려면 커서를 가져와야 한다
    c = conn.cursor()
    
    # 방법 1
    c.execute("DELETE FROM table1 WHERE id=?", (1,))
    # 방법 2
    c.execute("DELETE FROM table1 WHERE id=:id", {'id': 3})
    # 방법 3
    c.execute("DELETE FROM table1 WHERE id='%s'" % 5)
    # 확인
    for row in c.execute('SELECT * FROM table1'):
        print(row)
        
    conn.close()

def dbDump():   # Db 백업하기(dump.sql)
    conn = sqlite3.connect("test.db", isolation_level=None)
    # 커서 획득, 파이썬에서 파일을 읽고 쓰려면 커서를 가져와야 한다
    
    with conn:
        with open('dump.sql', 'w') as f:
            for line in conn.iterdump():
                f.write('%s\n' % line)
            print('Completed.')
                
    conn.close()
    
if __name__ == '__main__':
    print('----- dbInsert() -----')
    dbInsert()
    
    print('----- dbRead() -----')    
    dbRead()
    
    print('----- dbDump() -----')    
    dbDump()
    
    print('----- dbSearch() -----')    
    dbSearch()
    
    print('----- dbUpdate() -----')    
    dbUpdate()
    
    print('----- dbDelete() -----')    
    dbDelete() 

# %% 데이터베이스 활용(2) : sqlite3

import sqlite3

conn = sqlite3.connect('example.db')
#Creating a cursor object using the cursor() method
cursor = conn.cursor()
#Doping EMPLOYEE table if already exists.
cursor.execute("DROP TABLE IF EXISTS EMPLOYEE")

#Creating table as per requirement
sql ='''CREATE TABLE EMPLOYEE(
   FIRST_NAME CHAR(20) NOT NULL,
   LAST_NAME CHAR(20),
   AGE INT,
   SEX CHAR(1),
   INCOME FLOAT
)'''
cursor.execute(sql)
print("Table created successfully........")

# Commit your changes in the database
conn.commit()

#Closing the connection
conn.close()

try:
    sqliteConnection = sqlite3.connect('example.db')
    cursor = sqliteConnection.cursor()
    print("Successfully Connected to SQLite")

    insert_query = """INSERT INTO EMPLOYEE
                          (FIRST_NAME, LAST_NAME, AGE, SEX, INCOME) 
                           VALUES ('Gil Dong','Hong', 25, 1, 30000)"""

    count = cursor.execute(insert_query)
    sqliteConnection.commit()
    print("Record inserted successfully into SqliteDb_developers table ", cursor.rowcount)
    
    cursor.execute("SELECT * FROM EMPLOYEE")
    print(cursor.fetchall())    
    cursor.close()

except sqlite3.Error as error:
    print("Failed to insert data into sqlite table", error)
    
finally:
    if sqliteConnection:
        sqliteConnection.close()
        print("The SQLite connection is closed")
        
# %% 데이터베이스 활용(3) : contacts.db 생성
# QSqlDatabase /QSqlQuery /QTableWidget

import sys
from PyQt5.QtSql import QSqlDatabase
from PyQt5.QtSql import QSqlDatabase, QSqlQuery
from PyQt5.QtWidgets import QApplication, QMessageBox, QLabel

# Create the connection
con = QSqlDatabase.addDatabase("QSQLITE")
con.setDatabaseName("./contacts.db")

# Create the application
app = QApplication(sys.argv)

# Try to open the connection and handle possible errors
if not con.open():
    QMessageBox.critical(
        None,
        "App Name - Error!",
        "Database Error: %s" % con.lastError().databaseText(), )
    sys.exit(1)

# Create a query and execute it right away using .exec()
createTableQuery = QSqlQuery()
createTableQuery.exec(
    """
    CREATE TABLE contacts (
        id INTEGER PRIMARY KEY AUTOINCREMENT UNIQUE NOT NULL,
        name VARCHAR(40) NOT NULL,
        job VARCHAR(50),
        email VARCHAR(40) NOT NULL
    )
    """ )
print(con.tables())

# Creating a query for later execution using .prepare()
insertDataQuery = QSqlQuery()
insertDataQuery.prepare(
    """
    INSERT INTO contacts (
        name,
        job,
        email
    )
    VALUES (?, ?, ?)
    """ )

# Sample data
data = [
    ("Joe", "Senior Web Developer", "joe@example.com"),
    ("Lara", "Project Manager", "lara@example.com"),
    ("David", "Data Analyst", "david@example.com"),
    ("Jane", "Senior Python Developer", "jane@example.com"),]

# Use .addBindValue() to insert data
for name, job, email in data:
    insertDataQuery.addBindValue(name)
    insertDataQuery.addBindValue(job)
    insertDataQuery.addBindValue(email)
    insertDataQuery.exec()


# Create the application's window
win = QLabel(" Connection Successfully Opened! ")
win.setWindowTitle("App Name")
win.resize(250, 100)
win.show()
sys.exit(app.exec_())


# %% 데이터베이스 활용(3) : contacts.db 읽기
# QSqlDatabase /QSqlQuery /QTableWidget

import sys
from PyQt5.QtSql import QSqlDatabase, QSqlQuery
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QMessageBox,
    QTableWidget,
    QTableWidgetItem, )

class Contacts(QMainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("QTableView Example")
        self.resize(450, 250)
        # Set up the view and load the data
        self.view = QTableWidget()
        self.view.setColumnCount(4)
        self.view.setHorizontalHeaderLabels(["ID", "Name", "Job", "Email"])
        query = QSqlQuery("SELECT id, name, job, email FROM contacts")
        while query.next():
            rows = self.view.rowCount()
            self.view.setRowCount(rows + 1)
            self.view.setItem(rows, 0, QTableWidgetItem(str(query.value(0))))
            self.view.setItem(rows, 1, QTableWidgetItem(query.value(1)))
            self.view.setItem(rows, 2, QTableWidgetItem(query.value(2)))
            self.view.setItem(rows, 3, QTableWidgetItem(query.value(3)))
        self.view.resizeColumnsToContents()
        self.setCentralWidget(self.view )

def createConnection():
    con = QSqlDatabase.addDatabase("QSQLITE")
    con.setDatabaseName("./contacts.db")
    if not con.open():
        QMessageBox.critical(
            None,
            "QTableView Example - Error!",
            "Database Error: %s" % con.lastError().databaseText(),)
        return False
    return True

app = QApplication(sys.argv)
if not createConnection():
    sys.exit(1)
win = Contacts()
win.show()
sys.exit(app.exec_())

# %% 데이터베이스 활용(4) : QSqlDatabase /QSqlQuery /QSqlTableModel /setTable /QTableView

import sys
from PyQt5.QtSql import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *

def initializeModel(model):
    model.setTable('sportsmen')
    model.setEditStrategy(QSqlTableModel.OnFieldChange)
    model.select()
    model.setHeaderData(0, Qt.Horizontal, "ID")
    model.setHeaderData(1, Qt.Horizontal, "First name")
    model.setHeaderData(2, Qt.Horizontal, "Last name")

def createView(title, model):
    view = QTableView()
    view.setModel(model)
    view.setWindowTitle(title)
    return view

def addrow():
    print(model.rowCount())
    ret = model.insertRows(model.rowCount(), 1)
    print(ret)

def findrow(i):
    delrow = i.row()

def createDB():
    db = QSqlDatabase.addDatabase('QSQLITE')
    db.setDatabaseName('sports.db')

    if not db.open():
        QMessageBox.critical(None, qApp.tr("Cannot open database"),
                                   qApp.tr("Unable to establish a database connection.\n"
                                        "This example needs SQLite support. Please read "
                                        "the Qt SQL driver documentation for information "
                                        "how to build it.\n\n" "Click Cancel to exit."),
                                   QMessageBox.Cancel)
        return False

    query = QSqlQuery()
    query.exec_("create table sportsmen(id int primary key, "
                "firstname varchar(20), lastname varchar(20))")
    query.exec_("insert into sportsmen values(101, 'Roger', 'Federer')")
    query.exec_("insert into sportsmen values(102, 'Christiano', 'Ronaldo')")
    query.exec_("insert into sportsmen values(103, 'Ussain', 'Bolt')")
    query.exec_("insert into sportsmen values(104, 'Sachin', 'Tendulkar')")
    query.exec_("insert into sportsmen values(105, 'Saina', 'Nehwal')")
    db.close()
    return True

if __name__ == '__main__':
    app = QApplication(sys.argv)
    createDB()
    
    db = QSqlDatabase.addDatabase('QSQLITE')
    db.setDatabaseName('sports.db')
    
    model = QSqlTableModel()
    delrow = -1
    initializeModel(model)
    
    view1 = createView("Table Model (View 1)", model)
    view1.clicked.connect(findrow)
    
    dlg = QDialog()
    dlg.setGeometry(300,200,500,300)
    
    layout = QVBoxLayout()
    layout.addWidget(view1)
    
    button = QPushButton("Add a row")
    button.clicked.connect(addrow)
    layout.addWidget(button)
    
    btn1 = QPushButton("del a row")
    btn1.clicked.connect(lambda: model.removeRow(view1.currentIndex().row()))
    layout.addWidget(btn1)
    
    dlg.setLayout(layout)
    dlg.setWindowTitle("Database Demo")
    dlg.show()
    db.close()
    sys.exit(app.exec_())
    
# %% '학점 계산기' 소스 일부 - 실행 안됨 
# (https://ssola22.tistory.com/9), https://ssola22.tistory.com/8?category=548365)
# https://ssola22.tistory.com/category/Programming/PyQt
# https://ssola22.tistory.com/10?category=548365
# help(QtCore)

class Ui_load(object):
    
    def setupUi(self, load):
        load.resize(217, 157)
        load.setMinimumSize(QtCore.QSize(217, 157))                    #위젯 창의 최소 크기를 설정합니다
        load.setMaximumSize(QtCore.QSize(217, 157))                   #위젯 창의 최대 크기를 설정합니다
        load.setWindowIcon(QtGui.QIcon("english_ime-128.ico"))     #위젯의 윈도우 창에 있는 아이콘을 설정합니다
        self.lineEdit = QtGui.QLineEdit(load)                                #LineEdit을 불러옵니다
        self.lineEdit.setGeometry(QtCore.QRect(80, 20, 113, 20))       #LineEdit의 크기와 위치를 설정합니다        
        self.lineEdit.setMaxLength(8)                                          #최대 입력할 수 있는 텍스트 길이를 설정합니다
        self.lineEdit2 = QtGui.QLineEdit(load)
        self.lineEdit2.setGeometry(QtCore.QRect(80, 50, 113, 20))
        self.lineEdit2.setMaxLength(4)
        self.label_4 = QtGui.QLabel(load)
        self.label_4.setGeometry(QtCore.QRect(20, 20, 51, 20))
        self.label_4.setAlignment(QtCore.Qt.AlignCenter)            #Label의 글자를 중간으로 배열합니다
        self.label_6 = QtGui.QLabel(load)
        self.label_6.setGeometry(QtCore.QRect(20, 50, 51, 20))
        self.label_6.setAlignment(QtCore.Qt.AlignCenter)
        self.pushButton = QtGui.QPushButton(load)
        self.pushButton.setGeometry(QtCore.QRect(130, 120, 75, 23))
        self.label_5 = QtGui.QLabel(load)
        self.label_5.setGeometry(QtCore.QRect(10, 80, 201, 30)) 
        
        self.retranslateUi(load)
        QtCore.QObject.connect(self.pushButton, QtCore.SIGNAL("clicked()"), self.start)
        QtCore.QMetaObject.connectSlotsByName(load)
        
    def retranslateUi(self, load):
        load.setWindowTitle("안녕하세요!")
        self.label_4.setText("학번")
        self.label_6.setText("비밀번호")

    def start(self):
        global database
        num=self.lineEdit.text()
        pw=self.lineEdit2.text()
        c = mysql.connector.connect(user='root', password='apmsetup',host='000.000.000.000',database='semple')
        cursor = c.cursor()
        cursor.execute("select * from login where 학번=%s"%num)
        confirm = cursor.fetchone()
        
def table():
    
    cursor.execute("select * from ex2 where 학년=1 and 학기=1")
    result = cursor.fetchall()
    self.TableWidget.setColumnCount(4)        #세로 열 개수를 설정합니다
    self.TableWidget.setRowCount(12)           #가로 행 개수를 설정합니다. 적당한 행과 열이 존재하지않을 경우
    n=-1                                                     #테이블의 데이터가 보이지 않습니다!
    for i in result:
        n=n+1
        obje = QtGui.QTableWidgetItem(i[3])            #테이블의 데이터는 아이템 형식으로 입력합니다
        area = QtGui.QTableWidgetItem(i[2])            #DB의 데이터를 리스트로 잘라 순서대로 아이템으로 만듭니다
        grad = QtGui.QTableWidgetItem(i[5])
        score = QtGui.QTableWidgetItem(str(i[4]))
        self.all_Table.setItem(n, 1, obje)                  #아이템은 엑셀의 좌표와 같은 위치로 지정하여 입력합니다
        self.all_Table.setItem(n, 0, area)                  #즉 테이블의 제일 처음 셀의 위치는 (0,0)이 됩니다
        self.all_Table.setItem(n, 3, grad)
        self.all_Table.setItem(n, 2, score)

    ''' 테이블의 머리부분의 이름을 지정하는 메소드입니다. 머리 부분이므로 순서대로 인덱스가 정해집니다.'''
    Area = QtGui.QTableWidgetItem("영역");self.all_Table.setHorizontalHeaderItem(0, Area)    
    Obje = QtGui.QTableWidgetItem("과목");self.all_Table.setHorizontalHeaderItem(1, Obje)    
    Score= QtGui.QTableWidgetItem("학점");self.all_Table.setHorizontalHeaderItem(2, Score) 
    Grad = QtGui.QTableWidgetItem("성적");self.all_Table.setHorizontalHeaderItem(3, Grad)   
    self.all_Table.setColumnWidth(0,50);self.all_Table.setColumnWidth(1,180);    #가로 폭의 크기를 설정합니다
    self.all_Table.setColumnWidth(2,50);self.all_Table.setColumnWidth(3,50)      #해당 인덱스로 열을 선택합니다
    
    self.pushButton.setText( "시작하기")
    self.label_5.setText("※자신의 학번과 비밀번호 4자리를\n   입력해주십시오.")

    #ComboBox-----------------------------
    ''' 콤보박스도 아이템으로 데이터를 입력합니다. 아래와 같이 addItems매소드를 이용하여 리스트를 추가합니다.
        넣으려는 리스트 순서대로 콤보박스에 입력되어지며 사용자가 선택한 아이템의 인덱스를 받아들여서 사용합니다.'''
        
    scorecombo.addItems(["A+","A","B+","B","C+","C","D+","D","F","P"])
    
    ''' 만약 현재 B+을 선택한 상태이면 currentIndex메소드가 2라는 인덱스를 가져오고 itemText가 인덱스가 2에 해당하는
        텍스트 아이템을 가져와 변수 T에 할당하게됩니다.'''
        
    T=self.scorecombo.itemText(self.scorecombo.currentIndex())       
        
    #Listwidget-----------------------------
    ''' 리스트 위젯도 아이템으로 데이터를 입력하며 콤보박스와 원리가 똑같습니다.'''
    Listwidget.addItems(["항목1","항목2","항목3","항목4"])        
    Listwidget.addItem(i[3])                #기존에 존재하던 리스트의 항목도 추가할 수 있습니다.
    Listwidget.currentItem().text()        #리스트에서는 currentItem메소드를 이용해 
                                                     #아이템을 text화 하여 변수에 할당할 수 있습니다.

    #RadioButton---------------------------
    ''' 라디오버튼는 다수의 클릭 버튼 중에서 단 하나만을 체크해야 할 경우 사용됩니다. 중복하여 체크가 되지않죠!
        isChecked를 이용하면 해당하는 버튼이 체크되어있는 상태인지를 True / False로 반환합니다.
        체크가 되어있다면 True값을 반환합니다. 주로 if & else구문과 같이 이용하여 사용하게 됩니다.'''
        
    if self.RadioButton_1.isChecked():
        cursor.execute("select * from ex2 where 학년=1 and 학기=%s"%T)
        result = cursor.fetchall()
        for i in result:
             self.major_list.addItem(i[3])
    elif self.RadioButton_2.isChecked():
        
        '''{생략}'''

# %% csv 파일 쓰기

# 쓰기 방식(1)
import csv
data = [['서울1',2,3],[4,5,6],[7,8,9],[10,11,12],[13,14,15],[16,17,18]] # 이차원 리스트 
with open('./sample1.csv','w', newline='') as f: 
    makewrite = csv.writer(f) 
    for value in data: 
        makewrite.writerow(value)

# 쓰기 방식(2)
import csv
with open('./sample2.csv','w',newline='') as f: 
    wt = csv.writer(f) 
    wt.writerows(data)

# %% csv 파일 읽기

import csv
# 읽기 방식(1)
with open('./sample1.csv','r') as f: 
    reader = csv.reader(f) 
    print(reader)       # reader 객체를 출력 
    print(type(reader)) # reader 객체의 type을 출력 
    print(dir(reader))  # reader 객체에서 사용가능한 메서드 목록 출력
    
    for txt in reader: 
        print(txt)
    print()   

# 읽기 방식(2) - delimiter 지정
with open('./sample1.csv','r') as f: 
    reader = csv.reader(f, delimiter = '|') 
    for txt in reader: 
        print(txt)
    print()

# 읽기 방식(3) - 딕셔너리 형식으로 읽기
with open('./sample1.csv','r') as f: 
    reader = csv.DictReader(f) 
    for c in reader: 
        print(c)
    print()        

# %% pandas로 쓰기

import pandas as pd 
xlsx = pd.read_excel('./sample1.xlsx')

xlsx.to_excel('./pan_sample2.xlsx', index = False) # index : 첫 열에 숫자 붙여주기 
xlsx.to_csv('./pan_sample2.csv',index = False, encoding='utf-8-sig') 
# utf-8로 저장시 한글 깨, 해결책 : 'euc-kr' 또는 'utf-8-sig'


# %% pandas로 읽기

import pandas as pd
xlsx = pd.read_excel('./sample1.xlsx')  # 옵션: sheet_name='시트명 or 숫자', header=숫자, skiprow=숫자 
print(xlsx)
print()
print(xlsx.head())      # 상위 데이터 5개 확인 
print()
print(xlsx.tail())      # 하위 데이터 5개 확인 
print()
print(xlsx.shape)       # 행, 열
# print(xlsx.info())
# print(xlsx.describe())

# %%
# Import module
import sqlite3
  
# Connecting to sqlite
conn = sqlite3.connect('employee.db')
  
# Creating a cursor object using the cursor() method
cursor = conn.cursor()

# 데이터베이스 파일에 해당 테이블 존재 유무 확인하기
sql = "SELECT count(*) FROM sqlite_master WHERE type ='table' and name = 'EMPLOYEE'"
cursor.execute(sql)
conn.commit()

if cursor.fetchone()[0] != 1:
    print('테이블(MPLOYEE)이 존재하지 않아서, 생성합니다.')
    # Creating table
    table = """CREATE TABLE EMPLOYEE(FIRST_NAME VARCHAR(255), 
    LAST_NAME VARCHAR(255),AGE int, SEX VARCHAR(255), INCOME int);"""
    cursor.execute(table)
  
# Queries to INSERT records.
cursor.execute(
    '''INSERT INTO EMPLOYEE(FIRST_NAME, LAST_NAME, AGE, SEX, INCOME) 
    VALUES ('Anand', 'Choubey', 25, 'M', 10000)''')
  
cursor.execute(
    '''INSERT INTO EMPLOYEE(FIRST_NAME, LAST_NAME, AGE, SEX, INCOME) 
    VALUES ('Mukesh', 'Sharma', 20, 'M', 9000)''')
  
cursor.execute(
    '''INSERT INTO EMPLOYEE(FIRST_NAME, LAST_NAME, AGE, SEX, INCOME)
    VALUES ('Ankit', 'Pandey', 24, 'M', 6300)''')
  
cursor.execute(
    '''INSERT INTO EMPLOYEE(FIRST_NAME, LAST_NAME, AGE, SEX, INCOME)
    VALUES ('Subhdra ', 'Singh', 26, 'F', 8000)''')
  
cursor.execute(
    '''INSERT INTO EMPLOYEE(FIRST_NAME, LAST_NAME, AGE, SEX, INCOME)
    VALUES ('Tanu', 'Mishra', 24, 'F', 6500)''')
  
# Display data inserted
print("EMPLOYEE Table: ")
data = cursor.execute('''SELECT * FROM EMPLOYEE''')
for row in data:
    print(row)
  
# Updating
cursor.execute('''UPDATE EMPLOYEE SET INCOME = 5000 WHERE Age<25;''')
print('\nAfter Updating...\n')
  
# Display data
print("EMPLOYEE Table: ")
data = cursor.execute('''SELECT * FROM EMPLOYEE''')
for row in data:
    print(row)
  
# Commit your changes in the database
conn.commit()
  # Closing the connection
conn.close()

# %% 엑셀(*.xls) 자료 불러오기
import xlrd
import os, sys

path = os.getcwd()
print(path)

loc = "C:\\Users\\Administrator\\Documents\\DataDB\\DATA_2021_00.xlsx"

wb = xlrd.open_workbook(loc)
sheet = wb.sheet_by_index(0)
sheet.cell_value(0, 0)
 
for i in range(sheet.nrows):
    print(sheet.cell_value(i, 0))

# %% 엑셀(*.xlsx) 자료 불러오기
import openpyxl
import os, sys
from pathlib import Path

loc = "C:\\Users\\Administrator\\Documents\\DataDB\\"
file = "DATA_2021_00.xlsx"

#xlsx_file = loc + file
xlsx_file = Path(loc, file)

wb = openpyxl.load_workbook(xlsx_file) 
sheet = wb.active
# Read the active sheet:

max_row = sheet.max_row
max_column = sheet.max_column

print(max_row, max_column)

for row in sheet.iter_rows(max_row = 20):
    for col in row:
        print(col.value, end=" ")
    print()

# col_names = []
# for column in sheet.iter_cols(1, sheet.max_column):
#     col_names.append(column[0].value)

# print(col_names)

            